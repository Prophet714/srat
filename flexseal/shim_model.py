"""
Shim Model (SRAT-X)
-------------------
Reads shim + elastomer parameters from the Flexseal Design Tool
and computes shim stiffness contribution (K_shim).
"""

import openpyxl
from typing import Dict, Any
import numpy as np


def load_shim_params(
    excel_path: str,
    sheet_name: str = "Design sheet",
    cell_range: str = "D89:D96",
    label_range: str = "B89:B96",
) -> Dict[str, Any]:
    """
    Load shim + elastomer parameters from the Flexseal Design Tool.

    Parameters
    ----------
    excel_path : str
        Path to the Flexseal workbook (.xlsm).
    sheet_name : str
        Sheet containing shim/elastomer inputs.
    cell_range : str
        Range of input values (e.g., "D89:D96").
    label_range : str
        Range of parameter labels (e.g., "B89:B96").

    Returns
    -------
    dict
        {parameter_name: value}
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found in {excel_path}")

    ws = wb[sheet_name]

    values = [c[0].value for c in ws[cell_range]]
    labels = [c[0].value for c in ws[label_range]]

    results = {}
    for lbl, val in zip(labels, values):
        if lbl is None:
            continue
        key = str(lbl).strip().replace(" ", "_").lower()
        results[key] = val
    return results


def compute_kshim(params: Dict[str, Any]) -> float:
    """
    Compute shim rotational stiffness (placeholder model).

    Parameters
    ----------
    params : dict
        Dictionary from load_shim_params()

    Returns
    -------
    float
        Shim stiffness in lbf-in/deg
    """
    # Example: use E_s (shim hoop modulus) and nu (Poisson's ratio)
    # Placeholder until true physics from Flexseal Design Tool equations is added
    E_s = float(params.get("shim_hoop_modulus_of_elasticity", 0))
    G_e = float(params.get("elastomer_shear_modulus", 0))
    nu = float(params.get("elastomer_poisson's_ratio", 0))

    if E_s <= 0 or G_e <= 0:
        raise ValueError("Missing critical shim/elastomer parameters.")

    # Simplified placeholder stiffness model
    # In reality: tie to pad geometry and shim constraints
    k_lin = E_s / (1 - nu**2)  # compressive-like effective stiffness
    k_rot = k_lin * (np.pi / 180.0)  # convert to lbf-in/deg scale

    return k_rot


def compute_shim_stiffness(excel_path: str) -> float:
    """Convenience wrapper: load → compute."""
    params = load_shim_params(excel_path)
    return compute_kshim(params)


if __name__ == "__main__":
    path = r"C:\Users\U317688\OneDrive - L3Harris - GCCHigh\Sentinel\flexseal\Data\Sentinel Flexseal Preliminary Design Tool_v1.9.xlsm"
    k_shim = compute_shim_stiffness(path)
    print(f"K_shim = {k_shim:.2f} lbf-in/deg")
